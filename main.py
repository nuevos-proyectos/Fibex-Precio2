import os
import secrets
from fastapi import FastAPI, Depends, HTTPException, status, Form, Query, Request
from fastapi.responses import Response, JSONResponse, HTMLResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, EmailStr
from supabase import create_client, Client
from datetime import datetime
from dotenv import load_dotenv
from xhtml2pdf import pisa
from io import BytesIO
from typing import Optional
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import base64
import re  # Importante para limpiar letras en versiones antiguas

# Cargar variables de entorno ANTES de crear app
load_dotenv()

# Crear la app
app = FastAPI(title="Core Tabuladores Fibex Telecom")

# Montar archivos estáticos (debe ir después de crear app)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Configurar templates
templates = Jinja2Templates(directory="templates")

# Credenciales de conexión optimizadas
SUPABASE_URL = "https://muinkxeekvcivhtkiqvg.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im11aW5reGVla3ZjaXZodGtpcXZnIiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzk5ODMwNTcsImV4cCI6MjA5NTU1OTA1N30.exEvGrJervtUpa8muKjwImx-LVTnimhdmjJNVhl1mGE"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# --- ESQUEMAS DE DATOS ---
class UsuarioCreate(BaseModel):
    email: str
    password: str
    rol: int
    pais_asignado_id: Optional[int] = None
    region_asignada_id: Optional[int] = None
    nombre: Optional[str] = None
    apellido: Optional[str] = None
    departamento: Optional[str] = None

class PartidaCreate(BaseModel):
    id: Optional[int] = None
    categoria_id: int
    descripcion: str
    unidad: str

# --- SEGURIDAD: CONTROL DE SESIÓN Y ROLES ---
def resolver_usuario_por_token(token: str):
    res = supabase.table("usuarios").select("*").eq("token_sesion", token).execute()
    if not res.data:
        raise HTTPException(status_code=401, detail="Sesión inválida o expirada")
    return res.data[0]

# --- FUNCIÓN AUXILIAR DE AUDITORÍA ---
def registrar_auditoria(usuario_id: int, usuario_email: str, modulo: str, accion: str, descripcion: str):
    try:
        supabase.table("historial_cambios").insert({
            "usuario_id": usuario_id,
            "usuario_nombre": usuario_email,
            "modulo": modulo,
            "accion": accion,
            "descripcion": descripcion
        }).execute()
    except Exception as e:
        print(f"Error registrando auditoría: {e}")
        pass # Evitamos que un error en el log tumbe el proceso principal

@app.post("/auth/login")
def login(email: str = Form(...), password: str = Form(...)):
    res = supabase.from_("usuarios").select("*").eq("email", email).eq("password", password).execute()
    if not res.data:
        return JSONResponse(status_code=400, content={"error": "Credenciales incorrectas"})
    
    usuario = res.data[0]
    nuevo_token = secrets.token_hex(32)
    supabase.table("usuarios").update({"token_sesion": nuevo_token}).eq("id", usuario["id"]).execute()
    
    del usuario["password"]  # Seguridad
    
    # OPCIONAL: Registrar login en auditoría
    registrar_auditoria(usuario["id"], usuario["email"], "sistema", "ACCESO", "Usuario inició sesión")
    
    return {"token": nuevo_token, "perfil": usuario}

# --- METADATOS COMPLEMENTARIOS ---
@app.get("/data/auxiliar")
def obtener_auxiliares(token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    
    paises = supabase.table("paises").select("*").order("id").execute().data
    revisiones = supabase.table("revisiones").select("*").order("id", desc=True).execute().data
    categorias = supabase.table("categorias").select("*").order("orden_impresion").execute().data
    partidas = supabase.table("partidas").select("*").order("id").execute().data
    
    if profile["rol"] == 1:
        region_id = profile.get("region_asignada_id")
        regiones = []
        if region_id is not None:
            region_asignada = supabase.table("regiones").select("*").eq("id", region_id).single().execute().data
            if region_asignada:
                regiones.append(region_asignada)
                nacional = supabase.table("regiones").select("*").eq("pais_id", region_asignada["pais_id"]).eq("es_nacional", True).single().execute().data
                if nacional and nacional["id"] != region_asignada["id"]:
                    regiones.append(nacional)
    else:
        regiones = supabase.table("regiones").select("*").order("id").execute().data
        
    return {
        "paises": paises, 
        "regiones": regiones, 
        "revisiones": revisiones, 
        "categorias": categorias, 
        "partidas": partidas,
        "usuario_rol": profile["rol"]
    }

@app.get("/", response_class=HTMLResponse)
def leer_raiz(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

# --- GENERADOR PDF MATRICIAL DINÁMICO ---
@app.get("/tabulador/ver-pdf")
def ver_pdf(
    request: Request,
    token: str = Query(...),
    col_a_id: Optional[str] = Query(None),
    col_b_id: Optional[str] = Query(None),
    revision_id: Optional[str] = Query(None),
    mostrar_np: bool = Query(True),
    mostrar_om1: bool = Query(False),
    mostrar_om2: bool = Query(False)
):
    if col_a_id == "" or col_a_id == "none": col_a_id = None
    if col_b_id == "" or col_b_id == "none": col_b_id = None

    if revision_id == "": revision_id = None
    else:
        try: revision_id = int(revision_id) if revision_id else None
        except ValueError: revision_id = None

    profile = resolver_usuario_por_token(token)

    # --- lógica para rol 1 (permite elegir entre su región y el precio nacional) ---
    if profile["rol"] == 1:
        mostrar_om1 = False
        mostrar_om2 = False
        mostrar_np = True
        ciudad_id = profile.get("region_asignada_id")
        if not ciudad_id: raise HTTPException(status_code=400, detail="Usuario de rol 1 sin ciudad asignada.")
        ciudad_data = supabase.table("regiones").select("*").eq("id", ciudad_id).single().execute().data
        if not ciudad_data: raise HTTPException(status_code=404, detail="Región asignada a usuario no encontrada.")
        pais_id_ciudad = ciudad_data["pais_id"]
        nacional_data = supabase.table("regiones").select("*").eq("pais_id", pais_id_ciudad).eq("es_nacional", True).single().execute().data
        
        allowed_ids = {ciudad_id}
        if nacional_data:
            allowed_ids.add(nacional_data["id"])

        # Validar y asignar columna A
        if col_a_id is None:
            col_a_id = ciudad_id
        else:
            try:
                col_a_int = int(col_a_id)
                if col_a_int in allowed_ids:
                    col_a_id = col_a_int
                else:
                    col_a_id = ciudad_id
            except (ValueError, TypeError):
                col_a_id = ciudad_id

        # Validar y asignar columna B
        if col_b_id is not None and col_b_id != "none" and col_b_id != "":
            try:
                col_b_int = int(col_b_id)
                if col_b_int in allowed_ids and col_b_int != col_a_id:
                    col_b_id = str(col_b_int)
                else:
                    col_b_id = None
            except (ValueError, TypeError):
                col_b_id = None
        else:
            col_b_id = None

        rev_query = supabase.table("revisiones").select("id, numero_revision, fecha_creacion")\
            .eq("estado", "publicada").order("id", desc=True).limit(1).execute()
    else:
        if revision_id is None:
            rev_query = supabase.table("revisiones").select("id, numero_revision, fecha_creacion")\
                .eq("estado", "publicada").order("id", desc=True).limit(1).execute()
        else:
            rev_query = supabase.table("revisiones").select("id, numero_revision, fecha_creacion").eq("id", revision_id).execute()

    if not rev_query.data:
        raise HTTPException(status_code=404, detail="No hay tabulador activo/publicado.")
    
    revision = rev_query.data[0]
    revision_id_actual = revision["id"]

    try: col_a_id = int(col_a_id)
    except (ValueError, TypeError): raise HTTPException(status_code=400, detail="Columna A inválida.")

    region_a_data = supabase.table("regiones").select("*").eq("id", col_a_id).single().execute().data
    if not region_a_data: raise HTTPException(status_code=404, detail="Región A no encontrada.")
    pais_a_data = supabase.table("paises").select("moneda_simbolo").eq("id", region_a_data["pais_id"]).single().execute().data
    moneda_a = pais_a_data.get("moneda_simbolo", "$") if pais_a_data else "$"
    nombre_col_a = region_a_data["nombre"]

    print(f"DEBUG PDF INPUTS: col_a_id={col_a_id}, col_b_id={col_b_id}")
    comparar = False
    nombre_col_b = ""
    moneda_b = ""
    dict_b = {}
    
    if col_b_id and col_b_id != "none" and col_b_id != "":
        try:
            b_id = int(col_b_id)
            region_b_data = supabase.table("regiones").select("*").eq("id", b_id).single().execute().data
            if region_b_data:
                pais_b_data = supabase.table("paises").select("moneda_simbolo").eq("id", region_b_data["pais_id"]).single().execute().data
                moneda_b = pais_b_data.get("moneda_simbolo", "$") if pais_b_data else "$"
                nombre_col_b = region_b_data["nombre"]
                # Solo NP para la región B
                precios_b = supabase.table("precios").select("*").eq("revision_id", revision_id_actual).eq("region_id", b_id).execute().data
                dict_b = {p["partida_id"]: p["monto_usd"] for p in precios_b}
                comparar = True
                print(f"DEBUG comparar set to True for b_id={b_id}")
        except (ValueError, TypeError) as e:
            print(f"DEBUG exception in b_id processing: {e}")
            pass
    print(f"DEBUG FINAL comparar={comparar}")

    # Cargar NP para la región A
    precios_a = supabase.table("precios").select("*").eq("revision_id", revision_id_actual).eq("region_id", col_a_id).execute().data
    dict_a = {p["partida_id"]: p["monto_usd"] for p in precios_a}

    # Cargar O&M globales desde la nueva tabla
    precios_om = supabase.table("precios_om").select("*").eq("revision_id", revision_id_actual).execute().data
    om_dict = {p["partida_id"]: {"om1": p["monto_om_1"], "om2": p["monto_om_2"]} for p in precios_om}

    categorias_data = supabase.table("categorias").select("*").order("orden_impresion").execute().data or []
    partidas_data = supabase.table("partidas").select("*").order("id").execute().data or []

    documento_estructurado = []
    for cat in categorias_data:
        items_cat = []
        for part in partidas_data:
            if part["categoria_id"] == cat["id"]:
                precio_a_np = dict_a.get(part["id"], 0.0)
                precio_b_np = dict_b.get(part["id"], 0.0) if comparar else 0.0
                om = om_dict.get(part["id"], {"om1": 0.0, "om2": 0.0})
                items_cat.append({
                    "descripcion": part["descripcion"],
                    "unidad": part["unidad"],
                    "precio_a_np": precio_a_np,
                    "precio_b_np": precio_b_np,
                    "precio_a_om1": om["om1"],
                    "precio_a_om2": om["om2"],
                    # Si quieres mostrar comparativa de O&M (opcional), usa los mismos valores
                    "precio_b_om1": om["om1"] if comparar else 0.0,
                    "precio_b_om2": om["om2"] if comparar else 0.0,
                })
        if items_cat:
            documento_estructurado.append({"categoria": cat["nombre"], "partidas": items_cat})

    subtitulo_modo = " - COMPARATIVO" if comparar else ""
    tipos_precios = []
    if mostrar_np: tipos_precios.append("NP")
    if mostrar_om1: tipos_precios.append("O&M 1")
    if mostrar_om2: tipos_precios.append("O&M 2")
    
    fecha_creacion = revision.get("fecha_creacion")
    try: fecha_obj = datetime.fromisoformat(fecha_creacion) if isinstance(fecha_creacion, str) else fecha_creacion
    except Exception: fecha_obj = datetime.now()
    fecha_format = fecha_obj.strftime("%d/%m/%Y")

    logo_path = os.path.join("static", "logo.png")
    logo_base64 = ""
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as image_file: logo_base64 = base64.b64encode(image_file.read()).decode("utf-8")

    ctx = {
        "request": request,
        "revision": f"{revision['numero_revision']}{subtitulo_modo}",
        "fecha": fecha_format,
        "col_a_nombre": nombre_col_a,
        "col_b_nombre": nombre_col_b,
        "moneda_a": moneda_a,
        "moneda_b": moneda_b,
        "comparar": comparar,
        "mostrar_np": mostrar_np,
        "mostrar_om1": mostrar_om1,
        "mostrar_om2": mostrar_om2,
        "tipos_precios": tipos_precios,
        "datos": documento_estructurado,
        "logo_base64": logo_base64
    }

    html_content = templates.TemplateResponse("tabulador.html", ctx).body.decode("utf-8")
    pdf_file = BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=pdf_file)
    if pisa_status.err: raise HTTPException(status_code=500, detail="Error generando PDF.")

    return Response(content=pdf_file.getvalue(), media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="Tabulador_REV-{revision["numero_revision"]}.pdf"'})
# --- CAPA TOTAL DE ADMINISTRACIÓN (NIVEL 3 EXCLUSIVO) ---

# --- CAPA TOTAL DE ADMINISTRACIÓN ---

@app.post("/auth/cambiar-clave")
def cambiar_clave(token: str = Query(...), antigua_clave: str = Form(...), nueva_clave: str = Form(...)):
    profile = resolver_usuario_por_token(token)
    
    try:
        user_db = supabase.table("usuarios").select("password").eq("id", profile["id"]).single().execute().data
    except Exception:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    if not user_db or user_db["password"] != antigua_clave:
        raise HTTPException(status_code=400, detail="La contraseña anterior es incorrecta")
        
    supabase.table("usuarios").update({"password": nueva_clave}).eq("id", profile["id"]).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "usuarios", "MODIFICAR", "El usuario cambió su contraseña")
    return {"status": "Contraseña actualizada exitosamente"}

@app.post("/admin/usuarios")
def crear_usuario_global(user: UsuarioCreate, token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔒 EXCLUSIVO ROL 4: Solo el Súper Admin crea usuarios
    if profile["rol"] != 4: raise HTTPException(status_code=403, detail="No tienes permisos de Súper Administrador")
    
    user_data = user.model_dump()
    clean_data = {k: v for k, v in user_data.items() if v is not None}
    supabase.table("usuarios").insert(clean_data).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "usuarios", "CREAR", f"Creó un nuevo usuario: {user.email}")
    return {"status": "Usuario creado exitosamente"}

@app.get("/admin/usuarios")
def obtener_usuarios(token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔒 EXCLUSIVO ROL 4: Solo el Súper Admin (Rol 4)
    if profile["rol"] != 4:
        raise HTTPException(status_code=403, detail="No tienes permisos de Súper Administrador")
        
    usuarios_res = supabase.table("usuarios").select("*").order("email").execute().data or []
    regiones = supabase.table("regiones").select("id, nombre").execute().data or []
    paises = supabase.table("paises").select("id, nombre").execute().data or []
    
    region_map = {r["id"]: r["nombre"] for r in regiones}
    pais_map = {p["id"]: p["nombre"] for p in paises}
    
    for u in usuarios_res:
        u.pop("password", None)  # Eliminar clave por seguridad
        u["region_nombre"] = region_map.get(u.get("region_asignada_id"), "-")
        u["pais_nombre"] = pais_map.get(u.get("pais_asignado_id"), "-")
        
    return {"usuarios": usuarios_res}

@app.get("/admin/usuarios/exportar")
def exportar_usuarios_excel(token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔒 EXCLUSIVO ROL 4: Solo el Súper Admin (Rol 4)
    if profile["rol"] != 4:
        raise HTTPException(status_code=403, detail="No tienes permisos de Súper Administrador")
        
    usuarios_res = supabase.table("usuarios").select("*").order("email").execute().data or []
    regiones = supabase.table("regiones").select("id, nombre").execute().data or []
    paises = supabase.table("paises").select("id, nombre").execute().data or []
    
    region_map = {r["id"]: r["nombre"] for r in regiones}
    pais_map = {p["id"]: p["nombre"] for p in paises}
    
    roles_nombres = {
        1: "Nivel 1: Visualizador Regional",
        2: "Nivel 2: Visualizador General",
        3: "Nivel 3: Gerente de Nuevos Proyectos",
        4: "Nivel 4: Súper Administrador",
        5: "Nivel 5: Gerente de O&M"
    }

    # Crear libro de trabajo openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios Registrados"
    
    # Habilitar líneas de cuadrícula visibles
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos corporativos
    font_titulo = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    font_meta = Font(name="Calibri", size=10, bold=True, color="374151")
    font_meta_val = Font(name="Calibri", size=10, color="1F2937")
    font_headers = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=10, color="1F2937")
    
    fill_titulo = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
    fill_headers = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    fill_alt = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # 1. Título merged A1:F2
    ws.merge_cells("A1:F2")
    cell_titulo = ws["A1"]
    cell_titulo.value = "FIBEX TELECOM - REPORTE DE USUARIOS REGISTRADOS"
    cell_titulo.font = font_titulo
    cell_titulo.fill = fill_titulo
    cell_titulo.alignment = align_center
    
    # Rellenar las celdas unidas para evitar bordes blancos
    for r in range(1, 3):
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill_titulo
            
    # 2. Bloque de Metadatos
    fecha_hoy = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    ws["A4"] = "Fecha de Emisión:"
    ws["A4"].font = font_meta
    ws["B4"] = fecha_hoy
    ws["B4"].font = font_meta_val
    
    ws["A5"] = "Generado por:"
    ws["A5"].font = font_meta
    ws["B5"] = profile["email"]
    ws["B5"].font = font_meta_val
    
    # 3. Cabeceras de la tabla (Fila 7)
    headers = ["Nombre", "Apellido", "Departamento", "Correo Electrónico", "Rol de Usuario", "Ubicación Asignada"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_idx, value=text)
        cell.font = font_headers
        cell.fill = fill_headers
        cell.alignment = align_center
        cell.border = thin_border
        
    # 4. Filas de datos
    for row_idx, u in enumerate(usuarios_res, start=8):
        rol_id = u.get("rol", 0)
        
        # Determinar ubicación
        if rol_id == 1:
            ubicacion_str = region_map.get(u.get("region_asignada_id"), "-")
        else:
            ubicacion_str = "No aplica"
            
        rol_str = roles_nombres.get(rol_id, f"Rol {rol_id}")
        
        row_values = [
            u.get("nombre") or "-",
            u.get("apellido") or "-",
            u.get("departamento") or "-",
            u.get("email") or "-",
            rol_str,
            ubicacion_str
        ]
        
        is_even = (row_idx % 2 == 0)
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = thin_border
            
            # Alineación
            if col_idx in [5, 6]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
            # Cebra
            if is_even:
                cell.fill = fill_alt

    # 5. Ancho autoajustable para las columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"Reporte_Usuarios_Fibex_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.delete("/admin/usuarios/{usuario_id}")
def eliminar_usuario(usuario_id: int, token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔒 EXCLUSIVO ROL 4: Solo el Súper Admin (Rol 4)
    if profile["rol"] != 4:
        raise HTTPException(status_code=403, detail="No tienes permisos de Súper Administrador")
        
    # Evitar que el admin se borre a sí mismo
    if usuario_id == profile["id"]:
        raise HTTPException(status_code=400, detail="No puedes eliminar tu propio usuario administrador")
        
    try:
        usuario = supabase.table("usuarios").select("email").eq("id", usuario_id).single().execute().data
    except Exception:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
        
    # Eliminar usuario
    supabase.table("usuarios").delete().eq("id", usuario_id).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "usuarios", "ELIMINAR", f"Eliminó el usuario: {usuario['email']}")
    return {"status": "Usuario eliminado exitosamente"}

@app.post("/admin/partidas")
def crear_partida(partida: PartidaCreate, token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # ✅ PERMITIR ACCESO A 3, 4 y 5
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    
    data = partida.model_dump()
    if data.get("id") is None: data.pop("id")
    supabase.table("partidas").insert(data).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "partidas", "CREAR", f"Indexó nueva partida: {partida.descripcion}")
    return {"status": "Partida indexada"}
    
@app.put("/admin/partidas/{partida_id}")
def editar_partida(partida_id: int, partida: PartidaCreate, token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # ✅ PERMITIR ACCESO A 3, 4 y 5
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    
    supabase.table("partidas").update({
        "categoria_id": partida.categoria_id,
        "descripcion": partida.descripcion,
        "unidad": partida.unidad
    }).eq("id", partida_id).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "partidas", "MODIFICAR", f"Editó la partida ID {partida_id}: {partida.descripcion}")
    return {"status": "Partida actualizada"}

@app.delete("/admin/partidas/{partida_id}")
def borrar_partida(partida_id: int, token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔒 EXCLUSIVO ROL 4: Solo el Súper Admin puede borrar partidas
    if profile["rol"] != 4: 
        raise HTTPException(status_code=403, detail="No tienes permisos de Súper Administrador")
    
    try:
        partida = supabase.table("partidas").select("descripcion").eq("id", partida_id).single().execute().data
    except Exception:
        raise HTTPException(status_code=404, detail="Partida no encontrada")
    
    # Eliminar precios asociados en la tabla precios y precios_om
    supabase.table("precios").delete().eq("partida_id", partida_id).execute()
    supabase.table("precios_om").delete().eq("partida_id", partida_id).execute()
    
    # Eliminar la partida
    supabase.table("partidas").delete().eq("id", partida_id).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "partidas", "ELIMINAR", f"Eliminó la partida: {partida['descripcion']}")
    return {"status": "Partida eliminada exitosamente"}

@app.post("/admin/regiones")
def crear_region(token: str = Query(...), pais_id: int = Form(...), nombre: str = Form(...), es_nacional: bool = Form(False)):
    profile = resolver_usuario_por_token(token)
    # ✅ PERMITIR ACCESO A 3, 4 y 5
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    
    res = supabase.table("regiones").insert({"pais_id": pais_id, "nombre": nombre, "es_nacional": es_nacional}).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "regiones", "CREAR", f"Creó región: {nombre}")
    return {"status": "Región creada", "data": res.data}

@app.post("/admin/paises")
def crear_pais(token: str = Query(...), nombre: str = Form(...)):
    profile = resolver_usuario_por_token(token)
    # ✅ PERMITIR ACCESO A 3, 4 y 5
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    
    res_pais = supabase.table("paises").insert({"nombre": nombre}).execute().data[0]
    supabase.table("regiones").insert({"pais_id": res_pais["id"], "nombre": f"Nacional ({nombre})", "es_nacional": True}).execute()
    
    registrar_auditoria(profile["id"], profile["email"], "paises", "CREAR", f"Creó país: {nombre} y su base nacional")
    return {"status": "País creado con su base nacional"}

@app.get("/admin/precios/matriz")
def obtener_matriz_precios(token: str = Query(...), revision_id: int = Query(...), region_id: int = Query(...)):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [3,4,5]:
        raise HTTPException(status_code=403, detail="Permiso denegado")
    
    region_data = supabase.table("regiones").select("pais_id").eq("id", region_id).single().execute().data
    region_nacional_data = supabase.table("regiones").select("id").eq("pais_id", region_data["pais_id"]).eq("es_nacional", True).single().execute().data
    
    precios_regionales = supabase.table("precios").select("*").eq("revision_id", revision_id).eq("region_id", region_id).execute().data
    precios_nacionales = supabase.table("precios").select("*").eq("revision_id", revision_id).eq("region_id", region_nacional_data["id"]).execute().data
    
    # Obtener O&M globales
    precios_om = supabase.table("precios_om").select("*").eq("revision_id", revision_id).execute().data
    # Convertir a dict por partida_id
    om_dict = {p["partida_id"]: {"om1": p["monto_om_1"], "om2": p["monto_om_2"]} for p in precios_om}
    
    return {
        "nacionales": precios_nacionales,
        "regionales": precios_regionales,
        "precios_om": om_dict,
        "region_nacional_id": region_nacional_data["id"],
        "usuario_rol": profile["rol"]
    }

@app.post("/admin/precios/guardar")
def guardar_precios_simultaneos(
    token: str = Query(...), 
    revision_id: int = Form(...), 
    partida_id: int = Form(...), 
    region_id: int = Form(...), 
    monto_usd: float = Form(0.0),
    monto_om_1: float = Form(0.0),
    monto_om_2: float = Form(0.0)
):
    profile = resolver_usuario_por_token(token)
    user_rol = profile["rol"]
    if user_rol not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    
    rev_res = supabase.table("revisiones").select("estado").eq("id", revision_id).single().execute()
    if not rev_res.data or rev_res.data["estado"] != "borrador":
        raise HTTPException(status_code=400, detail="Solo puedes alterar precios de un borrador.")
    
    # 🧠 LÓGICA DE SALVAGUARDA DE COLUMNAS SEGÚN ROL:
    # Primero buscamos si ya existe un registro de precio para esa celda
    precio_existente = supabase.table("precios")\
        .select("*")\
        .eq("revision_id", revision_id)\
        .eq("partida_id", partida_id)\
        .eq("region_id", region_id)\
        .execute().data

    # Estructuramos el payload base
    payload = {
        "revision_id": revision_id,
        "partida_id": partida_id,
        "region_id": region_id
    }

    if precio_existente:
        registro_actual = precio_existente[0]
        if user_rol == 3:
            # Rol 3 (Nuevos Proyectos): Modifica su precio y mantiene congelados los O&M existentes
            payload["monto_usd"] = monto_usd
            payload["monto_om_1"] = registro_actual.get("monto_om_1", 0.0)
            payload["monto_om_2"] = registro_actual.get("monto_om_2", 0.0)
            tipo_cambio = "Nuevos Proyectos"
        elif user_rol == 5:
            # Rol 5 (O&M): Modifica O&M y mantiene congelado el monto base de Nuevos Proyectos
            payload["monto_usd"] = registro_actual.get("monto_usd", 0.0)
            payload["monto_om_1"] = monto_om_1
            payload["monto_om_2"] = monto_om_2
            tipo_cambio = "O&M 1 y 2"
        else:
            # Rol 4 (Súper Admin): Modifica absolutamente todo lo enviado
            payload["monto_usd"] = monto_usd
            payload["monto_om_1"] = monto_om_1
            payload["monto_om_2"] = monto_om_2
            tipo_cambio = "Completo (Admin)"
    else:
        # Si el registro no existe en la BD (fila totalmente nueva), insertamos según lo permitido
        payload["monto_usd"] = monto_usd if user_rol in [3, 4] else 0.0
        payload["monto_om_1"] = monto_om_1 if user_rol in [5, 4] else 0.0
        payload["monto_om_2"] = monto_om_2 if user_rol in [5, 4] else 0.0
        tipo_cambio = "Nuevo Registro"

    # Insertamos o actualizamos en Supabase
    supabase.table("precios").upsert(payload, on_conflict="revision_id,partida_id,region_id").execute()
    
    # AUDITORÍA
    try:
        nom_partida = supabase.table("partidas").select("descripcion").eq("id", partida_id).single().execute().data
        nom_region = supabase.table("regiones").select("nombre").eq("id", region_id).single().execute().data
        desc = nom_partida["descripcion"] if nom_partida else f"Partida {partida_id}"
        reg = nom_region["nombre"] if nom_region else f"Región {region_id}"
        
        mensaje = f"[{tipo_cambio}] Actualizó precios de '{desc}' en {reg}. Valores guardados -> USD: {payload['monto_usd']}, OM1: {payload['monto_om_1']}, OM2: {payload['monto_om_2']}"
        registrar_auditoria(profile["id"], profile["email"], "precios", "MODIFICAR", mensaje)
    except: pass
    
    return {"status": "Precio consolidado con éxito"}

@app.post("/admin/revisiones/crear-borrador")
def crear_borrador(token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [3, 4, 5]:
        raise HTTPException(status_code=403, detail="Permiso denegado")
    
    duplicados = supabase.table("revisiones").select("*").eq("estado", "borrador").execute().data
    if duplicados:
        raise HTTPException(status_code=400, detail="Ya existe un borrador abierto.")

    max_rev = supabase.table("revisiones").select("numero_revision").order("id", desc=True).limit(1).execute()
    nueva_rev_codigo = "1"
    if max_rev.data:
        try:
            nueva_rev_codigo = str(int(max_rev.data[0]["numero_revision"]) + 1)
        except ValueError:
            match = re.search(r'\d+', max_rev.data[0]["numero_revision"])
            nueva_rev_codigo = str(int(match.group()) + 1) if match else str(int(max_rev.data[0].get("id", 0)) + 1)

    ultima_pub = supabase.table("revisiones").select("*").eq("estado", "publicada").order("id", desc=True).limit(1).execute().data
    nueva_rev = supabase.table("revisiones").insert({"numero_revision": nueva_rev_codigo, "estado": "borrador"}).execute().data[0]
    
    if ultima_pub:
        # Clonar precios NP (tabla precios)
        precios_viejos = supabase.table("precios").select("*").eq("revision_id", ultima_pub[0]["id"]).execute().data
        if precios_viejos:
            clonados = [{"revision_id": nueva_rev["id"],
                         "partida_id": p["partida_id"],
                         "region_id": p["region_id"],
                         "monto_usd": p["monto_usd"],
                         "monto_om_1": p.get("monto_om_1", 0.0),
                         "monto_om_2": p.get("monto_om_2", 0.0)}
                        for p in precios_viejos]
            supabase.table("precios").insert(clonados).execute()
        
        # Clonar precios O&M (tabla precios_om)
        om_viejos = supabase.table("precios_om").select("*").eq("revision_id", ultima_pub[0]["id"]).execute().data
        if om_viejos:
            clonados_om = [{"revision_id": nueva_rev["id"],
                            "partida_id": o["partida_id"],
                            "monto_om_1": o["monto_om_1"],
                            "monto_om_2": o["monto_om_2"]}
                           for o in om_viejos]
            supabase.table("precios_om").insert(clonados_om).execute()
    
    # AUDITORÍA
    registrar_auditoria(profile["id"], profile["email"], "revisiones", "CREAR",
                        f"Creó borrador de revisión REV-{nueva_rev_codigo} y clonó matriz de precios")
    return {"status": "Borrador inicializado con datos clonados.", "nueva_version": nueva_rev_codigo}
    
@app.post("/admin/revisiones/publicar")
def publicar_revision(token: str = Query(...), revision_id: int = Form(...)):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")

    rev_res = supabase.table("revisiones").select("estado").eq("id", revision_id).single().execute()
    if not rev_res.data or rev_res.data["estado"] != "borrador": raise HTTPException(status_code=400, detail="Solo un borrador activo puede ser publicado.")
    
    supabase.table("revisiones").update({"estado": "archivada"}).eq("estado", "publicada").execute()
    supabase.table("revisiones").update({"estado": "publicada"}).eq("id", revision_id).execute()
    
    # AUDITORÍA
    registrar_auditoria(profile["id"], profile["email"], "revisiones", "MODIFICAR", f"Publicó de forma oficial la Revisión ID {revision_id}")
    return {"status": "Tabulador publicado oficialmente."}

@app.post("/admin/revisiones/cancelar-borrador")
def cancelar_borrador(token: str = Query(...), revision_id: Optional[int] = Form(None)):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [3, 4, 5]: raise HTTPException(status_code=403, detail="Permiso denegado")
    if revision_id is not None:
        rev_res = supabase.table("revisiones").select("*").eq("id", revision_id).single().execute()
        revision = rev_res.data
    else:
        borrador_res = supabase.table("revisiones").select("*").eq("estado", "borrador").order("id", desc=True).limit(1).execute()
        revision = borrador_res.data[0] if borrador_res.data else None

    if not revision: raise HTTPException(status_code=400, detail="No hay borrador activo para cancelar.")
    if revision.get("estado") != "borrador": raise HTTPException(status_code=400, detail="Solo un borrador puede ser cancelado.")

    supabase.table("precios").delete().eq("revision_id", revision["id"]).execute()
    supabase.table("revisiones").delete().eq("id", revision["id"]).execute()
    
    # AUDITORÍA
    registrar_auditoria(profile["id"], profile["email"], "revisiones", "ELIMINAR", f"Canceló y descartó el borrador de Revisión ID {revision['id']}")
    return {"status": "Borrador cancelado y eliminado correctamente."}

def logout(): pass

# --- NUEVO: ENDPOINTS DE AUDITORÍA ---
@app.get("/admin/auditoria/fechas")
def obtener_fechas_con_cambios(token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔴 CAMBIO AQUÍ: Ahora exige Nivel 4
    if profile["rol"] != 4: raise HTTPException(status_code=403, detail="Solo admin")
    
    res = supabase.table("historial_cambios").select("fecha_dia").order("fecha_dia", desc=True).execute()
    fechas = sorted(list(set([r["fecha_dia"] for r in res.data])), reverse=True)
    return {"fechas": fechas}

@app.get("/admin/auditoria/por-dia")
def obtener_cambios_del_dia(fecha: str = Query(...), token: str = Query(...)):
    profile = resolver_usuario_por_token(token)
    # 🔴 CAMBIO AQUÍ: Ahora exige Nivel 4
    if profile["rol"] != 4: raise HTTPException(status_code=403, detail="Solo admin")
    
    res = supabase.table("historial_cambios")\
        .select("*")\
        .eq("fecha_dia", fecha)\
        .order("fecha_hora", desc=True)\
        .execute()
    return res.data


# Esquema para O&M
class PrecioOM(BaseModel):
    revision_id: int
    partida_id: int
    monto_om_1: float = 0.0
    monto_om_2: float = 0.0

@app.get("/admin/precios/om")
def obtener_precios_om(token: str = Query(...), revision_id: int = Query(...)):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [3,4,5]:
        raise HTTPException(status_code=403, detail="Permiso denegado")
    
    res = supabase.table("precios_om").select("*").eq("revision_id", revision_id).execute()
    return {"precios": res.data}

@app.post("/admin/precios/om/guardar")
def guardar_precio_om(
    token: str = Query(...),
    revision_id: int = Form(...),
    partida_id: int = Form(...),
    monto_om_1: float = Form(0.0),
    monto_om_2: float = Form(0.0)
):
    profile = resolver_usuario_por_token(token)
    if profile["rol"] not in [4,5]:  # Solo admin o gerente O&M
        raise HTTPException(status_code=403, detail="Permiso denegado")
    
    # Verificar que la revisión sea borrador
    rev = supabase.table("revisiones").select("estado").eq("id", revision_id).single().execute()
    if not rev.data or rev.data["estado"] != "borrador":
        raise HTTPException(status_code=400, detail="Solo se puede editar en borrador")
    
    # Upsert
    data = {"revision_id": revision_id, "partida_id": partida_id, "monto_om_1": monto_om_1, "monto_om_2": monto_om_2}
    supabase.table("precios_om").upsert(data, on_conflict="revision_id,partida_id").execute()
    
    # Auditoría
    registrar_auditoria(profile["id"], profile["email"], "precios_om", "MODIFICAR", 
                        f"Guardó O&M para partida {partida_id}: OM1={monto_om_1}, OM2={monto_om_2}")
    return {"status": "ok"}

if __name__ == "__main__":
    import uvicorn
    import os
    # Python lee la variable de entorno directamente del sistema operativo
    puerto = int(os.environ.get("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=puerto)